"""Ocado Tuesday automation script."""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Literal

import openpyxl
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright


ROOT = Path(__file__).parent
XLSX_PATH = ROOT / "Ocado_Order_Manager.xlsx"
JSON_PATH = ROOT / "ocado_tuesday.json"  # canonical destination — newest export is copied here
DOWNLOADS_DIR = Path.home() / "Downloads"
# Match any export the checklist app produces: ocado_tuesday.json, ocado_tuesday30jun2026.json,
# ocado_tuesday_23jun.json, etc. Case-insensitive on the "ocado" stem.
JSON_GLOB = "[Oo]cado*.json"
SESSION_DIR = ROOT / ".ocado_session"
LOGS_DIR = ROOT / "logs"

URL_HOME = "https://www.ocado.com/"
# Use the home page as the anchor for login checks + as a known-good page to search from.
URL_RESERVED = URL_HOME
# Modern orders flow:
#   /orders                       → list of upcoming Wednesday cards
#   /orders/{ID}/details          → click into a specific week
#   then click "Edit order"       → enters editable basket
URL_ORDERS_LIST = "https://www.ocado.com/orders"
LOGIN_URL_FRAGMENTS = ("/login", "/signin", "sign-in", "log-in", "accounts.ocado", "/auth")
SKIP_CLEAR = False  # Set False once we've smoke-tested the clear-basket flow

# Selectors — comma-separated fallbacks. Tweak on first run if Ocado has changed.
SEL_SEARCH_INPUT = (
    'input[placeholder*="Find a product" i], '
    'input[placeholder*="Find" i], '
    'input[aria-label*="search" i], '
    'input[role="searchbox"], '
    'input[role="combobox"], '
    '[role="search"] input, '
    'input[type="search"], '
    'input[name="search"], '
    '[data-testid="search-input"]'
)
SEL_SEARCH_SUBMIT = 'button[type="submit"][aria-label*="earch"], button:has-text("Search")'
SEL_PRODUCT_CARD = (
    'a[href*="/products/"], '
    '[data-testid*="product-tile"] a, '
    '[data-testid*="product-card"] a, '
    'article[data-product-id] a, '
    'li[data-product-id] a, '
    '[data-sku] a'
)
SEL_PRODUCT_TITLE = 'h1, [data-testid="product-title"]'
SEL_ADD_BUTTON = 'button:has-text("Add"), button[data-testid*="add-to-trolley"]'
SEL_QTY_PLUS = 'button[aria-label*="ncrease"], button:has-text("+")'
SEL_REMOVE_ALL = 'button:has-text("Remove all"), button:has-text("Empty trolley"), button:has-text("Clear trolley")'
SEL_REMOVE_CONFIRM = 'button:has-text("Yes"), button:has-text("Confirm"), button:has-text("Remove")'
SEL_BASKET_BTN = (
    '[data-test="basket-button"], [data-synthetics="basket-details-button"], '
    'button[aria-label*="basket" i], button[aria-label*="trolley" i], '
    'a[aria-label*="basket" i], a[aria-label*="trolley" i]'
)
SEL_CHECKOUT_BTN = (
    '[data-test="basket-checkout-button"], [data-synthetics="start-checkout-button"], '
    'button:has-text("Check out to save changes"), '
    'a:has-text("Check out"), button:has-text("Check out"), '
    'a:has-text("Checkout"), button:has-text("Checkout")'
)
SEL_CONTINUE_BTN = (
    '[data-synthetics="next-step"], '
    'a:has-text("Continue checkout"), button:has-text("Continue checkout"), '
    'a:has-text("Continue"), button:has-text("Continue")'
)
SEL_PLACE_ORDER_BTN = (
    '[data-test="place-order-button"], [data-synthetics="place-order-button"], '
    'button:has-text("Place order"), button:has-text("Place Order")'
)
SEL_ORDER_CONFIRMED = (
    'h1:has-text("Order confirmation"), h1:has-text("Order Confirmation"), '
    ':has-text("Your order has been placed"), :has-text("Order confirmed")'
)
URL_ORDER_CONFIRMED_PATTERNS = ("confirmation", "thank-you", "thankyou", "order-placed", "success")
# Positive logged-in markers — only appear when authenticated
SEL_LOGGED_IN_MARKER = 'a:has-text("Sign out"), a:has-text("Log out"), button:has-text("Sign out"), button:has-text("Log out")'
# Logged-out / login-page markers — fields and buttons only present when NOT authenticated
SEL_LOGGED_OUT_MARKER = (
    'input[type="password"], '
    'a:has-text("Sign in"), a:has-text("Log in"), '
    'button:has-text("Sign in"), button:has-text("Log in")'
)

ELEMENT_TIMEOUT_MS = 15_000
NAVIGATION_TIMEOUT_MS = 30_000


def setup_logging() -> Path:
    LOGS_DIR.mkdir(exist_ok=True)
    log_path = LOGS_DIR / f"run-{datetime.now():%Y-%m-%d-%H%M}.log"
    handler_file = logging.FileHandler(log_path)
    handler_stdout = logging.StreamHandler(sys.stdout)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    for h in (handler_file, handler_stdout):
        h.setFormatter(fmt)
    logging.basicConfig(level=logging.INFO, handlers=[handler_file, handler_stdout])
    return log_path


@dataclass
class Item:
    name: str
    qty: int
    search_term: str
    notes: str
    source: Literal["tier1", "tier2"]


@dataclass
class TrolleyItem:
    """A manually-added trolley item captured before the order is emptied."""
    name: str
    url: str
    qty: int


def load_tier1(xlsx_path: Path) -> list[Item]:
    """Load active rows from the 'Tier 1 – Essentials' sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tier 1 – Essentials"]
    items: list[Item] = []
    # Row 1: title; row 2: headers; data from row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, _category, qty, search_term, notes, active = (row + (None,) * 6)[:6]
        if not name:
            continue
        if str(active or "").strip().lower() != "yes":
            continue
        items.append(
            Item(
                name=str(name).strip(),
                qty=int(qty or 1),
                search_term=str(search_term or name).strip(),
                notes=str(notes or "").strip(),
                source="tier1",
            )
        )
    return items


def resolve_tier2_json_path(log: logging.Logger) -> Path | None:
    """Find the Tier 2 JSON — the newest 'ocado*.json' export.

    Scans both ~/Downloads and the project root for files matching JSON_GLOB
    (ocado_tuesday.json, ocado_tuesday30jun2026.json, ocado_tuesday_23jun.json, …)
    and picks the most recently modified. If the winner lives in Downloads (or is
    a differently-named root file), it's copied to the canonical JSON_PATH so the
    rest of the run has one known file. Returns None if no export is found.
    """
    candidates = list(DOWNLOADS_DIR.glob(JSON_GLOB)) + list(ROOT.glob(JSON_GLOB))
    if not candidates:
        return None

    newest = max(candidates, key=lambda p: p.stat().st_mtime)
    mtime = datetime.fromtimestamp(newest.stat().st_mtime)
    log.info("Newest Tier 2 export: %s (modified %s)", newest.name, f"{mtime:%Y-%m-%d %H:%M}")

    if newest.resolve() != JSON_PATH.resolve():
        import shutil
        try:
            src_display = newest.relative_to(Path.home())
        except ValueError:
            src_display = newest
        log.info("Copying ~/%s → %s", src_display, JSON_PATH.name)
        shutil.copy2(newest, JSON_PATH)

    return JSON_PATH


def upcoming_wednesday(today: date) -> date:
    """The Wednesday on or after `today`."""
    # Monday=0 ... Wednesday=2 ... Sunday=6
    delta = (2 - today.weekday()) % 7
    return today + timedelta(days=delta)


def load_tier2(json_path: Path) -> tuple[list[Item], str]:
    """Return (items, week_string). Raises FileNotFoundError if missing."""
    if not json_path.exists():
        raise FileNotFoundError(json_path)
    with json_path.open() as f:
        data = json.load(f)
    week_str = str(data.get("week", "")).strip()
    items: list[Item] = []
    for entry in data.get("tier2_yes", []):
        name = str(entry.get("name", "")).strip()
        if not name:
            continue
        items.append(
            Item(
                name=name,
                qty=int(entry.get("qty", 1)),
                search_term=str(entry.get("search_term") or name).strip(),
                notes=str(entry.get("notes") or "").strip(),
                source="tier2",
            )
        )
    return items, week_str


def login_state(page: Page) -> Literal["in", "out", "unknown"]:
    """Positive-marker login detection. Returns 'in' / 'out' / 'unknown'."""
    url = page.url
    if any(frag in url for frag in LOGIN_URL_FRAGMENTS):
        return "out"
    try:
        if page.locator(SEL_LOGGED_IN_MARKER).first.is_visible(timeout=1_500):
            return "in"
    except PWTimeout:
        pass
    try:
        if page.locator(SEL_LOGGED_OUT_MARKER).first.is_visible(timeout=1_500):
            return "out"
    except PWTimeout:
        pass
    return "unknown"


def ensure_logged_in(page: Page) -> None:
    log = logging.getLogger("ocado")
    page.goto(URL_RESERVED, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    try:
        page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
    except PWTimeout:
        pass
    for attempt in (1, 2):
        state = login_state(page)
        log.info("Login check: state=%s url=%s", state, page.url)
        if state == "in":
            return
        if state == "unknown":
            log.warning("Login state unclear — please confirm manually.")
        else:
            log.warning("Not logged in (attempt %d).", attempt)
            if _auto_login(page, log):
                page.goto(URL_RESERVED, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
                try:
                    page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
                except PWTimeout:
                    pass
                continue
        print(f">>> Browser is at: {page.url}")
        print(">>> If not already logged in, log in now. Then press Enter here to continue.")
        input(">>> ")
        page.goto(URL_RESERVED, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
        try:
            page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
        except PWTimeout:
            pass
    if login_state(page) == "out":
        raise RuntimeError("Still not logged in after retry — aborting.")


def _auto_login(page: Page, log: logging.Logger) -> bool:
    """Attempt to log in using credentials stored in macOS Keychain.

    Returns True if login was attempted, False if no credentials found.
    """
    try:
        import keyring
        email = keyring.get_password("ocado", "email")
        password = keyring.get_password("ocado", email or "")
        if not email or not password:
            return False
    except Exception as e:  # noqa: BLE001
        log.warning("Keychain lookup failed: %s", e)
        return False

    log.info("Credentials found in Keychain — attempting auto-login for %s", email)
    try:
        # Navigate to login if not already there
        if not any(f in page.url for f in LOGIN_URL_FRAGMENTS):
            page.goto("https://www.ocado.com/login", wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
        try:
            page.wait_for_load_state("networkidle", timeout=15_000)
        except PWTimeout:
            pass

        # Ocado SSO uses username/email field — try several selector patterns
        email_input = page.locator(
            'input[type="email"], input[name="email"], input[id*="email" i], '
            'input[name="username"], input[id*="username" i], input[id*="user" i]'
        ).first
        email_input.wait_for(state="visible", timeout=10_000)
        email_input.fill(email)

        # Some SSO flows show password on same page, others need a Next click first
        pw_input = page.locator('input[type="password"]').first
        try:
            pw_input.wait_for(state="visible", timeout=3_000)
        except PWTimeout:
            # Password not visible yet — click Next/Continue to advance to step 2
            page.locator('button[type="submit"], input[type="submit"], button:has-text("Next"), button:has-text("Continue")').first.click()
            pw_input.wait_for(state="visible", timeout=10_000)

        pw_input.fill(password)
        page.locator('button[type="submit"], input[type="submit"]').first.click()
        try:
            page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
        except PWTimeout:
            pass
        log.info("Auto-login submitted.")
        return True
    except PWTimeout as e:
        log.warning("Auto-login timed out: %s", e)
        return False


def scrape_order_urls(page: Page, log: logging.Logger) -> list[str]:
    """Collect unique product URLs from the current edit-order view."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15_000)
    except PWTimeout:
        pass
    links = page.locator('a[href*="/products/"]').all()
    seen: set[str] = set()
    urls: list[str] = []
    for link in links:
        href = link.get_attribute("href") or ""
        if href.startswith("/"):
            href = "https://www.ocado.com" + href
        base = href.split("?")[0]
        if base and "/products/" in base and base not in seen:
            seen.add(base)
            urls.append(base)
    log.info("Scraped %d product URL(s) from current order", len(urls))
    return urls


def readd_by_url(page: Page, url: str, log: logging.Logger) -> None:
    """Navigate to a product page by URL and click Add."""
    log.info("Re-adding pre-clear item: %s", url)
    dismiss_modals(page, log)
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
        add_btn = page.locator(SEL_ADD_BUTTON).first
        add_btn.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
        add_btn.click()
        log.info("   re-added OK")
    except PWTimeout as e:
        log.warning("   could not re-add %s: %s", url, e)


def _debug_screenshot(page: Page, log: logging.Logger, label: str) -> None:
    path = LOGS_DIR / f"debug-{datetime.now():%H%M%S}-{label}.png"
    try:
        page.screenshot(path=str(path))
        log.info("DEBUG screenshot: %s (url=%s)", path.name, page.url)
    except Exception as e:  # noqa: BLE001
        log.warning("DEBUG screenshot failed: %s", e)


def _debug_buttons(page: Page, log: logging.Logger, label: str) -> None:
    try:
        buttons = page.locator("button, [role='button']").all()
        texts = [((b.text_content() or "").strip()) for b in buttons if b.is_visible()]
        texts = [t for t in texts if t]
        log.info("DEBUG buttons at %s: %s", label, texts)
    except Exception as e:  # noqa: BLE001
        log.warning("DEBUG buttons failed: %s", e)


# Ocado's trolley drawer marks each line item with these attributes. Scoping to
# them is essential: the drawer also contains recommendation carousels, so a
# page-wide 'a[href*="/products/"]' scrape picks up ~4x too many items.
JS_READ_TROLLEY = """() => {
    const rows = [...document.querySelectorAll('[data-test="expanded-trolley-list-item"]')];
    return rows.map(row => {
        const links = [...row.querySelectorAll('a[href*="/products/"]')];
        let href = links.length ? (links[0].getAttribute('href') || '') : '';
        if (href.startsWith('/')) href = 'https://www.ocado.com' + href;
        // Each row has an image link (empty text) and a title link — take the longest text.
        const name = links
            .map(a => (a.textContent || '').trim())
            .sort((a, b) => b.length - a.length)[0] || '';
        const qtyEl = row.querySelector('[data-test="quantity-in-basket"]');
        return {
            href: href.split('?')[0],
            name,
            qtyRaw: qtyEl ? (qtyEl.textContent || '').trim() : '',
        };
    });
}"""


def capture_trolley_items(page: Page, log: logging.Logger) -> list[TrolleyItem]:
    """Snapshot manually-added trolley items before the order is emptied.

    Opens the trolley drawer, reads each line item (URL + quantity), then closes
    the drawer. Returns [] if the trolley is empty or unreadable — callers should
    treat that as "nothing to preserve" rather than an error.
    """
    log.info("Capturing current trolley contents…")
    try:
        page.locator(SEL_BASKET_BTN).first.click(timeout=10_000)
    except PWTimeout:
        log.warning("Basket button not found — cannot capture trolley items")
        return []

    # Wait for the drawer to actually render its rows before reading.
    try:
        page.locator('[data-test="expanded-trolley-list-item"]').first.wait_for(
            state="visible", timeout=8_000
        )
    except PWTimeout:
        log.info("   trolley appears empty — nothing to preserve")
        dismiss_modals(page, log)
        return []

    raw = page.evaluate(JS_READ_TROLLEY)

    items: list[TrolleyItem] = []
    for entry in raw:
        href = entry.get("href") or ""
        if not href or "/products/" not in href:
            log.warning("   skipping trolley row with no product link: %r", entry.get("name"))
            continue
        digits = re.search(r"\d+", entry.get("qtyRaw") or "")
        qty = int(digits.group()) if digits else 1
        items.append(TrolleyItem(name=entry.get("name") or href, url=href, qty=qty))
        log.info("   captured: %s x%d  (qty text: %r)",
                 items[-1].name[:50], qty, entry.get("qtyRaw"))

    log.info("Trolley snapshot: %d item(s) to restore after clearing", len(items))
    dismiss_modals(page, log)
    return items


def restore_trolley_items(page: Page, items: list[TrolleyItem], log: logging.Logger) -> None:
    """Re-add previously captured trolley items, skipping any already in the basket.

    Runs after the Tier 1/Tier 2 loop, so anything the lists already added is
    detected by its quantity counter and left alone rather than double-added.
    """
    if not items:
        return
    log.info("Restoring %d preserved trolley item(s)…", len(items))
    for item in items:
        log.info("→ restoring %s x%d", item.name[:50], item.qty)
        dismiss_modals(page, log)
        try:
            page.goto(item.url, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
            dismiss_modals(page, log)

            # A visible increment counter means it's already in the basket
            # (added by Tier 1/Tier 2) — leave it as-is.
            counter = page.locator('[data-test="counter:increment"]').first
            try:
                counter.wait_for(state="visible", timeout=2_500)
                log.info("   already in basket — skipping")
                continue
            except PWTimeout:
                pass

            add_btn = page.locator(SEL_ADD_BUTTON).first
            add_btn.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
            add_btn.click()
            log.info("   added 1")
            page.wait_for_timeout(1_500)
            dismiss_modals(page, log)

            for n in range(item.qty - 1):
                dismiss_modals(page, log)
                plus = page.locator(SEL_QTY_PLUS).first
                plus.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
                plus.click()
                log.info("   qty +1 (now %d)", n + 2)
                page.wait_for_timeout(500)
        except PWTimeout as e:
            log.warning("   could not restore %s: %s", item.name[:50], e)
        except Exception as e:  # noqa: BLE001
            log.warning("   could not restore %s: %s: %s", item.name[:50], type(e).__name__, e)


def clear_reserved_order(page: Page, target_date: date) -> None:
    """Navigate /orders → matching Wednesday card → Edit order → remove all.

    Ocado lists upcoming orders as cards with text like 'Wed 20 May 9:00pm - 10:00pm'.
    We match on 'Wed {DD} {Mon}' built from target_date.
    """
    log = logging.getLogger("ocado")
    log.info("Clearing reserved order for %s…", target_date.isoformat())
    page.goto(URL_ORDERS_LIST, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    try:
        page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
    except PWTimeout:
        pass

    month_str = target_date.strftime('%b')
    day_variants = {str(target_date.day), f"{target_date.day:02d}"}
    day_strs = [f"Wed {d} {month_str}" for d in day_variants]
    log.info("Looking for order card matching any of: %s", day_strs)
    card = page.locator(", ".join(f':has-text("{d}")' for d in day_strs)).last  # `.last` skips the page heading
    try:
        card.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
    except PWTimeout:
        log.warning("No order card found for any of %s — skipping clear.", day_strs)
        return
    card.click()
    try:
        page.wait_for_url("**/orders/*/details*", timeout=NAVIGATION_TIMEOUT_MS)
    except PWTimeout:
        log.warning("Did not reach /orders/*/details after click — skipping clear.")
        return

    _debug_screenshot(page, log, "after-card-click")

    edit_btn = page.locator('button:has-text("Edit order"), a:has-text("Edit order")').first
    try:
        edit_btn.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
        edit_btn.click()
        log.info("Clicked Edit order.")
    except PWTimeout:
        _debug_buttons(page, log, "no-edit-order-button")
        log.warning("No 'Edit order' button — basket may already be in edit mode or empty.")
    try:
        page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
    except PWTimeout:
        pass

    _debug_screenshot(page, log, "after-edit-order")
    _debug_buttons(page, log, "before-popup-handler")

    # Ocado shows a popup after clicking Edit order.
    # Affirmative buttons (proceed into edit mode): "Confirm", "Keep order"
    # Dismiss buttons (cancel editing):             "Not now", "Cancel order"
    # We always want the affirmative one.
    # Use get_by_role to catch both <button> tags and role="button" elements.
    clicked_popup = False
    for popup_label in ("Confirm", "Keep order", "Keep Order"):
        btn = page.get_by_role("button", name=popup_label, exact=True)
        try:
            btn.wait_for(state="visible", timeout=2_000)
            btn.click()
            log.info("Clicked edit-order popup: %r", popup_label)
            clicked_popup = True
            try:
                page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
            except PWTimeout:
                pass
            break
        except PWTimeout:
            continue

    if not clicked_popup:
        # Fallback: JS click — works regardless of element type
        js_clicked = page.evaluate("""() => {
            const labels = ['Confirm', 'Keep order', 'Keep Order'];
            const els = [...document.querySelectorAll('button, [role="button"], a')];
            for (const el of els) {
                const text = (el.textContent || '').trim();
                if (labels.includes(text)) {
                    el.click();
                    return text;
                }
            }
            return null;
        }""")
        if js_clicked:
            log.info("Clicked edit-order popup via JS: %r", js_clicked)
            clicked_popup = True
            try:
                page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
            except PWTimeout:
                pass
        else:
            log.info("No edit-order popup detected.")

    _debug_screenshot(page, log, "after-popup-handler")

    remove_btn = page.locator(SEL_REMOVE_ALL).first
    try:
        remove_btn.wait_for(state="visible", timeout=5_000)
        remove_btn.click()
        log.info("Clicked 'Remove all'.")
        # Confirmation dialog: "Yes, empty trolley" / "Yes" depending on Ocado version.
        # Use get_by_role first (catches <button> and role="button"), then JS fallback.
        confirmed = False
        for confirm_label in ("Yes, empty trolley", "Yes", "Confirm"):
            btn = page.get_by_role("button", name=confirm_label, exact=False)
            try:
                btn.first.wait_for(state="visible", timeout=3_000)
                btn.first.click()
                log.info("Confirmed removal: %r", confirm_label)
                confirmed = True
                break
            except PWTimeout:
                continue
        if not confirmed:
            js_clicked = page.evaluate("""() => {
                const labels = ['Yes, empty trolley', 'Yes, empty', 'Yes'];
                const els = [...document.querySelectorAll('button, [role="button"]')];
                for (const el of els) {
                    const text = (el.textContent || '').trim();
                    if (labels.some(l => text.startsWith(l))) {
                        el.click();
                        return text;
                    }
                }
                return null;
            }""")
            if js_clicked:
                log.info("Confirmed removal via JS: %r", js_clicked)
            else:
                log.info("No confirmation dialog — that's fine.")
    except PWTimeout:
        _debug_buttons(page, log, "no-remove-all-button")
        log.info("No 'Remove all' button found in edit view — basket may already be empty.")


@dataclass
class Result:
    item: Item
    status: Literal["ok", "not_found", "error"]
    detail: str = ""
    product_url: str = ""


# Matches sizes that confuse Ocado's search: '300g', '1.5l', '8x330ml', '6 x 90g', etc.
# Ocado returns zero results when the search term contains a specific size.
_SIZE_PATTERN = re.compile(
    r'\s*\b\d+(?:\.\d+)?\s*(?:[x×]\s*\d+(?:\.\d+)?\s*)?(?:g|kg|mg|ml|cl|l|oz|lb|pints?|pt)\b\s*',
    re.IGNORECASE,
)


def strip_size_suffix(term: str) -> str:
    """Remove '300g', '1kg', '750ml', '2 pints' etc. from a search term."""
    cleaned = _SIZE_PATTERN.sub(' ', term)
    return re.sub(r'\s+', ' ', cleaned).strip()


SEL_MODAL_CLOSE = (
    'button[aria-label="Close modal"], '
    'button[aria-label="Close"], '
    '.ReactModalPortal button[aria-label*="lose"], '
    '.ReactModalPortal button[class*="close"], '
    '.ReactModalPortal button[class*="__button--close"]'
)


def dismiss_modals(page: Page, log: logging.Logger) -> None:
    """Close any open modal/dialog overlay. Safe to call when none is present."""
    # Use JS to check if the portal has rendered any content — more reliable than
    # Playwright's is_visible() on the portal container, whose own layout may be empty.
    has_modal = page.evaluate("""() => {
        const portal = document.querySelector('.ReactModalPortal');
        return !!(portal && portal.children.length > 0);
    }""")
    if not has_modal:
        return

    log.info("   modal detected, attempting dismiss")

    # Try the × close button first (clean path)
    try:
        close = page.locator(SEL_MODAL_CLOSE).first
        close.wait_for(state="visible", timeout=2_000)
        close.click()
        log.info("   clicked modal close button")
        page.wait_for_timeout(300)
        return
    except PWTimeout:
        pass

    # Forcibly remove the ReactModalPortal from the DOM — works regardless of
    # whether the modal responds to Escape or backdrop clicks.
    removed = page.evaluate("""() => {
        const portal = document.querySelector('.ReactModalPortal');
        if (portal && portal.children.length > 0) {
            portal.innerHTML = '';
            return true;
        }
        return false;
    }""")
    if removed:
        log.info("   cleared ReactModalPortal via JS")
        page.wait_for_timeout(300)


def _try_search_suggestion(page: Page, log: logging.Logger) -> bool:
    """If page shows 'No results for', click the first suggestion chip and return True."""
    no_results = page.get_by_text("No results for", exact=False).first
    try:
        no_results.wait_for(state="visible", timeout=2_000)
    except PWTimeout:
        return False
    suggestion = no_results.locator('xpath=following::button[1]')
    try:
        suggestion.wait_for(state="visible", timeout=2_000)
    except PWTimeout:
        return False
    text = (suggestion.text_content() or "").strip()
    log.info("   no results — clicking suggestion: %r", text)
    suggestion.click()
    try:
        page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
    except PWTimeout:
        pass
    return True


def _titles_roughly_match(actual_title: str, search_term: str) -> bool:
    """True if the longest word in search_term appears (case-insensitive) in actual_title."""
    words = [w for w in search_term.split() if len(w) >= 4]
    if not words:
        return True
    longest = max(words, key=len)
    return longest.lower() in actual_title.lower()


def add_item(page: Page, item: Item, capture_only: bool = False) -> Result:
    """Search for or navigate to the item's product page and add it to the basket.

    If capture_only=True, navigate to the product page but do NOT add to basket
    or change quantity. Used to populate URL cache without touching the basket.
    """
    log = logging.getLogger("ocado")
    log.info("→ %s [%s] qty=%d%s", item.name, item.source, item.qty, " (capture-only)" if capture_only else "")
    dismiss_modals(page, log)
    try:
        if item.notes.startswith("https://www.ocado.com"):
            page.goto(item.notes, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
        else:
            search_term = strip_size_suffix(item.search_term)
            if search_term != item.search_term:
                log.info("   search term cleaned: '%s' → '%s'", item.search_term, search_term)
            search_box = page.locator(SEL_SEARCH_INPUT).first
            search_box.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
            search_box.fill(search_term)
            search_box.press("Enter")
            try:
                page.wait_for_load_state("networkidle", timeout=NAVIGATION_TIMEOUT_MS)
            except PWTimeout:
                pass
            first_card = page.locator(SEL_PRODUCT_CARD).first
            try:
                first_card.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
            except PWTimeout:
                # "No results" fallback: Ocado often shows a shorter suggestion chip
                # ("Here are some product recommendations…"). Try clicking the first
                # such chip and re-check for products.
                if _try_search_suggestion(page, log):
                    try:
                        first_card.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
                    except PWTimeout:
                        return Result(item, "not_found", f"no results for '{search_term}' (suggestion also empty)")
                else:
                    return Result(item, "not_found", f"no results for '{search_term}'")
            first_card.click()

        title_loc = page.locator(SEL_PRODUCT_TITLE).first
        title_loc.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
        actual = (title_loc.text_content() or "").strip()
        if not _titles_roughly_match(actual, item.search_term):
            log.warning("Title mismatch: wanted ~'%s', got '%s' — continuing", item.search_term, actual)

        if capture_only:
            log.info("   captured URL: %s", page.url)
            return Result(item, "ok", actual, product_url=page.url)

        add_btn = page.locator(SEL_ADD_BUTTON).first
        add_btn.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
        add_btn.click()
        log.info("   added 1")
        page.wait_for_timeout(1_500)  # give the favourites modal time to appear
        dismiss_modals(page, log)

        for n in range(item.qty - 1):
            dismiss_modals(page, log)  # modal may have re-appeared between qty clicks
            plus = page.locator(SEL_QTY_PLUS).first
            plus.wait_for(state="visible", timeout=ELEMENT_TIMEOUT_MS)
            plus.click()
            log.info("   qty +1 (now %d)", n + 2)
            page.wait_for_timeout(500)  # brief pause for any modal after qty click

        return Result(item, "ok", actual, product_url=page.url)
    except PWTimeout as e:
        return Result(item, "error", f"timeout: {e}")
    except Exception as e:  # noqa: BLE001
        return Result(item, "error", f"{type(e).__name__}: {e}")


def print_summary(results: list[Result], log_path: Path) -> None:
    tier1 = [r for r in results if r.item.source == "tier1"]
    tier2 = [r for r in results if r.item.source == "tier2"]
    ok1 = sum(1 for r in tier1 if r.status == "ok")
    ok2 = sum(1 for r in tier2 if r.status == "ok")
    not_found = [r for r in results if r.status == "not_found"]
    errors = [r for r in results if r.status == "error"]

    print()
    print("=" * 41)
    print("              SUMMARY")
    print("=" * 41)
    print(f"Tier 1: {ok1}/{len(tier1)} added")
    print(f"Tier 2: {ok2}/{len(tier2)} added")
    if not_found:
        print(f"\nNot found ({len(not_found)}):")
        for r in not_found:
            print(f"  - {r.item.name} [{r.item.source}]   search: \"{r.item.search_term}\"")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for r in errors:
            print(f"  - {r.item.name} [{r.item.source}]: {r.detail}")
    print(f"\nFull log: {log_path}")
    print("=" * 41)


def write_product_urls(xlsx_path: Path, results: list[Result], log: logging.Logger) -> int:
    """Write captured product URLs back to the spreadsheet.

    - Tier 1 results → "Tier 1 – Essentials" sheet, Notes column (col 5)
    - Tier 2 results → "Rotation Pool" sheet, Notes column (col 7)

    Matches by exact Item Name. Always overwrites existing Notes content for
    matching rows. Returns the number of cells updated across both sheets.
    """
    tier1_urls = {
        r.item.name: r.product_url
        for r in results
        if r.item.source == "tier1" and r.status == "ok" and r.product_url
    }
    tier2_urls = {
        r.item.name: r.product_url
        for r in results
        if r.item.source == "tier2" and r.status == "ok" and r.product_url
    }
    if not tier1_urls and not tier2_urls:
        log.info("No URLs to write back.")
        return 0
    wb = openpyxl.load_workbook(xlsx_path)
    updated = 0
    if tier1_urls:
        ws = wb["Tier 1 – Essentials"]
        # Notes is column 5 (index 4)
        for row in ws.iter_rows(min_row=3):
            name_cell = row[0]
            if not name_cell.value:
                continue
            name = str(name_cell.value).strip()
            if name in tier1_urls:
                row[4].value = tier1_urls[name]
                updated += 1
    if tier2_urls:
        ws = wb["Rotation Pool"]
        # Notes is column 7 (index 6)
        matched = set()
        for row in ws.iter_rows(min_row=3):
            name_cell = row[0]
            if not name_cell.value:
                continue
            name = str(name_cell.value).strip()
            if name in tier2_urls:
                row[6].value = tier2_urls[name]
                matched.add(name)
                updated += 1
        unmatched = set(tier2_urls) - matched
        if unmatched:
            log.warning(
                "Tier 2 names not found in Rotation Pool sheet (URL not written): %s",
                ", ".join(sorted(unmatched)),
            )
    wb.save(xlsx_path)
    return updated


def confirm_write_urls(count: int) -> bool:
    print(f"\n{count} product URL(s) captured this run.")
    answer = input("Write them back to the spreadsheet's Notes column? [y/N]: ").strip().lower()
    return answer == "y"


def confirm_stale_week(week_str: str, expected: date) -> bool:
    print(f"\n⚠  Tier 2 JSON week is '{week_str}', expected '{expected.isoformat()}'.")
    answer = input("   Proceed anyway? [y/N]: ").strip().lower()
    return answer == "y"


def confirm_checkout() -> bool:
    answer = input("\nProceed to checkout and place the order? [y/N]: ").strip().lower()
    return answer == "y"


def checkout(page: Page, log: logging.Logger) -> bool:
    """Click through the Ocado checkout flow and place the order.

    Returns True if the order confirmation page is reached.
    """
    log.info("Starting checkout flow")

    # Step 1: open the basket panel
    try:
        page.locator(SEL_BASKET_BTN).first.click()
        log.info("Clicked basket button")
    except PWTimeout:
        log.error("Checkout: basket button not found — aborting")
        return False

    # Step 2: click the primary checkout button in the panel
    try:
        page.locator(SEL_CHECKOUT_BTN).first.click(timeout=10_000)
        log.info("Clicked checkout button")
    except PWTimeout:
        log.error("Checkout: checkout button not found — aborting")
        return False

    # Step 3: click Continue until we see Place order (max 5 steps)
    for step in range(1, 6):
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except PWTimeout:
            pass

        place_order = page.locator(SEL_PLACE_ORDER_BTN).first
        try:
            place_order.wait_for(state="visible", timeout=3_000)
            log.info("Reached Place order button")
            break
        except PWTimeout:
            pass

        try:
            page.locator(SEL_CONTINUE_BTN).first.click(timeout=10_000)
            log.info("Checkout step %d: clicked Continue", step)
        except PWTimeout:
            log.warning("Checkout step %d: no Continue button found", step)
            break

    # Step 4: place the order
    try:
        page.locator(SEL_PLACE_ORDER_BTN).first.click(timeout=10_000)
        log.info("Clicked Place order")
    except PWTimeout:
        log.error("Checkout: Place order button not found — aborting")
        return False

    # Step 5: wait for confirmation page — detected by URL pattern or page content
    try:
        page.wait_for_load_state("domcontentloaded", timeout=30_000)
    except PWTimeout:
        pass
    url = page.url.lower()
    if any(p in url for p in URL_ORDER_CONFIRMED_PATTERNS):
        log.info("Order confirmation received! (url=%s)", page.url)
        return True
    try:
        page.locator(SEL_ORDER_CONFIRMED).first.wait_for(state="visible", timeout=10_000)
        log.info("Order confirmation received!")
        return True
    except PWTimeout:
        log.warning("Order confirmation page not detected — check the browser manually (url=%s)", page.url)
        return False


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ocado Tuesday automation")
    parser.add_argument(
        "--capture-only",
        action="store_true",
        help="Visit each Tier 1 item's product page to grab its URL and write it back to the spreadsheet's Notes column. Skips clearing the basket and skips adding items. Items that already have a URL in Notes are skipped.",
    )
    parser.add_argument(
        "--probe-trolley",
        action="store_true",
        help="Read and print the current trolley contents, then exit. Changes nothing — use this to check the capture is seeing the right items before a real run.",
    )
    return parser.parse_args(argv)


def probe_trolley(log: logging.Logger) -> int:
    """Read-only: print what capture_trolley_items() sees, then exit."""
    log.warning("PROBE-TROLLEY mode: reading the trolley only. Nothing will be changed.")
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(SESSION_DIR),
            headless=False,
            viewport={"width": 1400, "height": 900},
        )
        context.set_default_timeout(ELEMENT_TIMEOUT_MS)
        context.set_default_navigation_timeout(NAVIGATION_TIMEOUT_MS)
        page = context.pages[0] if context.pages else context.new_page()
        try:
            ensure_logged_in(page)
            items = capture_trolley_items(page, log)
            print()
            print("=" * 41)
            print(f"       TROLLEY: {len(items)} ITEM(S)")
            print("=" * 41)
            for it in items:
                print(f"  {it.qty} x  {it.name}")
                print(f"         {it.url}")
            print("=" * 41)
            print("Nothing was changed. If this list matches your trolley, the")
            print("real run will preserve exactly these items.")
        except Exception as e:  # noqa: BLE001
            log.exception("Probe failed: %s", e)
        print("\nBrowser left open. Close the Chromium window when done.")
        try:
            page.wait_for_event("close", timeout=0)
        except Exception:  # noqa: BLE001
            pass
        try:
            context.close()
        except Exception:  # noqa: BLE001
            pass
    return 0


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    log_path = setup_logging()
    log = logging.getLogger("ocado")
    log.info("Run starting. Log: %s", log_path)
    if args.capture_only:
        log.warning("CAPTURE-ONLY mode: visiting items to grab URLs. Basket will NOT be touched.")

    if args.probe_trolley:
        return probe_trolley(log)

    tier1 = load_tier1(XLSX_PATH)
    log.info("Loaded %d Tier 1 items", len(tier1))

    if args.capture_only:
        # In capture mode we only need Tier 1; skip items that already have a URL.
        items_to_visit = [
            it for it in tier1
            if not it.notes.startswith("https://www.ocado.com")
        ]
        skipped = len(tier1) - len(items_to_visit)
        log.info("Capturing URLs for %d items (%d already have a URL, skipped)", len(items_to_visit), skipped)
        expected_week = upcoming_wednesday(date.today())
    else:
        json_path = resolve_tier2_json_path(log)
        if json_path is None:
            log.error("No Tier 2 JSON (%s) found in %s or %s — aborting.", JSON_GLOB, DOWNLOADS_DIR, ROOT)
            return 1
        tier2, week_str = load_tier2(json_path)
        log.info("Loaded %d Tier 2 items (week=%s) from %s", len(tier2), week_str, json_path.name)

        expected_week = upcoming_wednesday(date.today())
        if week_str != expected_week.isoformat():
            if not confirm_stale_week(week_str, expected_week):
                log.info("User declined stale JSON — aborting.")
                return 1
        items_to_visit = tier1 + tier2

    all_items = items_to_visit
    results: list[Result] = []

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(SESSION_DIR),
            headless=False,
            viewport={"width": 1400, "height": 900},
        )
        context.set_default_timeout(ELEMENT_TIMEOUT_MS)
        context.set_default_navigation_timeout(NAVIGATION_TIMEOUT_MS)
        page = context.pages[0] if context.pages else context.new_page()

        preserved: list[TrolleyItem] = []
        try:
            ensure_logged_in(page)
            if args.capture_only:
                log.info("Skipping basket clear (capture-only mode).")
                page.goto(URL_HOME, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
            elif SKIP_CLEAR:
                log.warning("SKIP_CLEAR=True — basket-clearing disabled. Clear the basket manually in the browser, then press Enter.")
                input(">>> Press Enter once the basket is empty (or to proceed anyway): ")
                page.goto(URL_HOME, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
            else:
                # Emptying the order also wipes anything manually added to the
                # trolley, and Ocado gives no way to merge the two — so snapshot
                # the trolley now and replay it after the lists are added.
                preserved = capture_trolley_items(page, log)
                clear_reserved_order(page, expected_week)
            for i, item in enumerate(all_items, 1):
                result = add_item(page, item, capture_only=args.capture_only)
                results.append(result)
                if result.status == "ok":
                    log.info("[%d/%d] OK: %s", i, len(all_items), item.name)
                elif result.status == "not_found":
                    log.warning("[%d/%d] NOT FOUND: %s (%s)", i, len(all_items), item.name, result.detail)
                else:
                    log.error("[%d/%d] ERROR: %s (%s)", i, len(all_items), item.name, result.detail)

            # Replay the pre-clear trolley last, so items the lists already
            # added are detected and skipped rather than double-added.
            if not args.capture_only:
                restore_trolley_items(page, preserved, log)

        except Exception as e:  # noqa: BLE001
            log.exception("Run aborted: %s", e)

        print_summary(results, log_path)
        if preserved:
            print(f"\nPreserved trolley items re-added: {len(preserved)}")
            for it in preserved:
                print(f"  {it.qty} x {it.name[:60]}")

        capturable = sum(1 for r in results if r.status == "ok" and r.product_url)
        if capturable > 0 and confirm_write_urls(capturable):
            try:
                n = write_product_urls(XLSX_PATH, results, log)
                log.info("Wrote %d URLs to %s", n, XLSX_PATH.name)
            except PermissionError:
                log.error("Could not write to %s — is Excel/Numbers holding it open? Close it and try again.", XLSX_PATH.name)
            except Exception as e:  # noqa: BLE001
                log.exception("Failed to write URLs: %s", e)

        if not args.capture_only and confirm_checkout():
            checkout(page, log)

        print("\nBrowser left open for review. Close the Chromium window when done — your session will be saved.")
        try:
            page.wait_for_event("close", timeout=0)
        except Exception:  # noqa: BLE001
            pass
        try:
            context.close()
        except Exception:  # noqa: BLE001
            pass

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
