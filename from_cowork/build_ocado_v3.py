from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PURPLE = "6B3FA0"
GREEN = "2E7D32"
WHITE = "FFFFFF"
AMBER = "FFF8E1"
RED_LITE = "FFEBEE"
PINK = "FCE4EC"
HEADER_TXT = "FFFFFF"

CAT_COLOURS = {
    "Fruit": "E8F5E9", "Veg": "DCEDC8", "Dairy": "E3F2FD",
    "Cheese": "FFF9C4", "Eggs": "FFF3E0", "Breakfast": "FBE9E7",
    "Condiments": "F3E5F5", "Drinks": "E0F2F1", "Snacks": "FBE9E7",
    "Protein": "FCE4EC", "Household": "ECEFF1", "Health": "E8EAF6",
    "Frozen": "E1F5FE", "Pasta": "F3E5F5", "Treats": "FCE4EC",
}

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def style_cell(cell, bold=False, color=WHITE, size=10, align="left", wrap=False):
    cell.font = Font(name="Calibri", bold=bold, color="000000" if color != WHITE else "000000", size=size)
    cell.fill = fill(color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border

def write_banner(ws, row, text, col_count, bg=PURPLE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

def write_headers(ws, row, headers, bg=GREEN):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[row].height = 18

def write_data_row(ws, row, values, row_bg=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        needs_attention = isinstance(v, str) and v.startswith("⚠")
        if needs_attention:
            bg = PINK
        elif row_bg:
            bg = row_bg
        else:
            bg = CAT_COLOURS.get(values[1], "FFFFFF") if len(values) > 1 else "FFFFFF"
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 16

# ── TIER 1 DATA ──────────────────────────────────────────────────────────────
tier1_rows = [
    ("M&S Pink Lady Apples",               "Fruit",      1, "M&S Pink Lady Apples",                                 "",                                   "Yes"),
    ("M&S Organic Fairtrade Bananas",       "Fruit",      1, "M&S Organic Fairtrade Bananas",                        "",                                   "Yes"),
    ("M&S Blueberries Family Pack",         "Fruit",      2, "M&S Blueberries Family Pack 300g",                     "",                                   "Yes"),
    ("M&S Raspberries",                     "Fruit",      1, "M&S Raspberries 250g",                                 "",                                   "Yes"),
    ("M&S Blackberries",                    "Fruit",      1, "M&S Blackberries 150g",                                "",                                   "Yes"),
    ("M&S Seedless White Grapes",           "Fruit",      1, "M&S Seedless White Grapes 500g",                       "",                                   "Yes"),
    ("Oranges",                             "Fruit",      1, "M&S Easy Peelers",                                     "⚠ Confirm exact product/variety",    "Yes"),
    ("M&S Extra Fine Beans",                "Veg",        1, "M&S Extra Fine Beans 200g",                            "",                                   "Yes"),
    ("M&S Organic Broccoli",                "Veg",        1, "M&S Organic Broccoli 350g",                            "",                                   "Yes"),
    ("M&S Organic Cucumber",                "Veg",        1, "M&S Organic Cucumber",                                 "",                                   "Yes"),
    ("M&S Organic Peppers",                 "Veg",        1, "M&S Organic Peppers",                                  "",                                   "Yes"),
    ("M&S Organic Cherry Tomatoes",         "Veg",        1, "M&S Organic Piccolini Cherry Tomatoes Vine Ripened",   "",                                   "Yes"),
    ("Yeo Valley Organic Whole Milk",       "Dairy",      1, "Yeo Valley Organic Fresh Whole Milk 1L",               "",                                   "Yes"),
    ("Moo Long Life Whole Milk",            "Dairy",      1, "Moo Whole Long Life Milk 1L",                          "",                                   "Yes"),
    ("M&S Mature Grated Cheddar",           "Dairy",      1, "M&S British Mature Grated Cheddar 250g",               "",                                   "Yes"),
    ("M&S Double Cream",                    "Dairy",      1, "M&S British Double Cream 600ml",                       "",                                   "Yes"),
    ("Gouda",                               "Cheese",     1, "M&S Dutch Gouda Sliced",                               "⚠ Confirm exact product",            "Yes"),
    ("Burford Brown Large Eggs x10",        "Eggs",       3, "Clarence Court Burford Brown 10 Large Free Range Eggs","",                                   "Yes"),
    ("Cheerios Honey Multigrain",           "Breakfast",  1, "Cheerios Honey Multigrain Breakfast Cereal 370g",      "",                                   "Yes"),
    ("Jordans Strawberry Country Crisp",    "Breakfast",  1, "Jordans Country Crisp Sun-Ripe Strawberry 450g",       "",                                   "Yes"),
    ("Protein Everything Bagels",           "Breakfast",  1, "New York Bakery Co Protein Everything Bagels",         "⚠ Confirm exact product",            "Yes"),
    ("Oatly Barista",                       "Breakfast",  1, "Oatly Barista Edition",                                "",                                   "Yes"),
    ("Ketchup",                             "Condiments", 1, "Heinz Tomato Ketchup",                                 "⚠ Confirm brand/size preferred",     "Yes"),
    ("Mayo",                                "Condiments", 1, "Hellmann's Real Mayonnaise",                           "⚠ Confirm brand/size preferred",     "Yes"),
    ("Maple Syrup",                         "Condiments", 1, "Pure Canadian Maple Syrup",                            "⚠ Confirm brand preferred",          "Yes"),
    ("Diet Coke No Caffeine",               "Drinks",     1, "Diet Coke No Caffeine 8x330ml",                        "",                                   "Yes"),
    ("Bastides Petits Saucissons Secs",     "Snacks",     2, "Bastides Petits Saucissons Secs",                      "Buy 2 for £7 deal",                  "Yes"),
    ("Popz Butter Microwave Popcorn",       "Snacks",     1, "Popz Butter Microwave Popcorn 6x90g",                  "",                                   "Yes"),
    ("M&S Mint Crumbles",                   "Snacks",     2, "M&S Mint Crumbles 178g",                               "",                                   "Yes"),
]

# ── ROTATION POOL DATA ────────────────────────────────────────────────────────
# (Item, Category, Tier, Qty, Search Term, Last Ordered, Notes)
rotation_rows = [
    # Tier 2 — Protein
    ("M&S 2 Chicken Kyivs",               "Protein",   2, 2, "M&S 2 Chicken Kyivs 320g",                              "", "Rotate with other proteins"),
    ("M&S Organic Chicken Drumsticks",    "Protein",   2, 3, "M&S Organic British Chicken Drumsticks 500g",           "", "Rotate with other proteins"),
    ("M&S Duck Breast Portions",          "Protein",   2, 2, "M&S Select Farms British 2 Duck Breast Portions 265g",  "", "Buy 2 for £10 deal"),
    ("M&S Scottish Salmon Fillets",       "Protein",   2, 1, "M&S 4 Scottish Salmon Fillets Skin On 480g",            "", "Rotate with other proteins"),
    ("Unearthed Mild Kabanos",            "Protein",   2, 1, "Unearthed Mild Kabanos 105g",                           "", ""),
    ("Pork Chops",                        "Protein",   2, 1, "M&S British Pork Loin Steaks",                          "", "⚠ Confirm exact cut/product"),
    ("Turkey Steaks",                     "Protein",   2, 1, "M&S British Turkey Breast Steaks",                      "", "⚠ Confirm exact product"),
    # Tier 2 — Cheese
    ("Snowdonia Black Bomber Cheddar",    "Cheese",    2, 1, "Snowdonia Black Bomber Extra Mature Cheddar 200g",      "", ""),
    ("M&S Somerset Ripening Brie",        "Cheese",    2, 1, "M&S Somerset Ripening Brie 230g",                       "", ""),
    ("Boursin Garlic & Herbs",            "Cheese",    2, 1, "Boursin Garlic Herbs Soft French Cheese 150g",          "", ""),
    ("M&S Shaved Parmigiano Reggiano",    "Cheese",    2, 1, "M&S Shaved Parmigiano Reggiano 80g",                    "", ""),
    ("Cathedral City Mini Snack Cheeses", "Cheese",    2, 1, "Cathedral City Mini Mature Snack Cheeses 6x20g",        "", "Kids snack"),
    ("Cheestrings Original",              "Cheese",    2, 1, "Cheestrings Original 8x20g",                            "", "Kids snack"),
    ("Leerdammer",                        "Cheese",    2, 2, "Leerdammer Original 200g",                              "", ""),
    # Tier 2 — Dairy
    ("M&S Salted Butter from Brittany",   "Dairy",     2, 1, "M&S Salted Butter from Brittany 250g",                  "", ""),
    ("Yeo Valley Organic Unsalted Butter","Dairy",     2, 1, "Yeo Valley Organic Unsalted Butter 200g",               "", ""),
    # Tier 2 — Drinks
    ("Cawston Press Cloudy Apple Juice",  "Drinks",    2, 1, "Cawston Press Cloudy Apple Juice 1L",                   "", ""),
    ("Dalston's Sparkling Cherry",        "Drinks",    2, 1, "Dalston's Sparkling Cherry 4x330ml",                    "", ""),
    ("London Essence Pink Grapefruit",    "Drinks",    2, 1, "London Essence Co Pink Grapefruit Cans 6x150ml",        "", ""),
    ("Caffeinated Diet Coke / Pepsi Max", "Drinks",    2, 1, "Pepsi Max 8x330ml",                                     "", "⚠ Confirm: Pepsi Max or regular Diet Coke?"),
    # Tier 2 — Frozen
    ("Strong Roots Sweet Potato Fries",   "Frozen",    2, 1, "Strong Roots Oven Baked Sweet Potato Fries 500g",       "", ""),
    ("Aunt Bessie's Roast Potatoes",      "Frozen",    2, 1, "Aunt Bessie's Roast Potatoes 1.1kg",                    "", ""),
    # Tier 2 — Fruit
    ("M&S Seedless Easy Peelers",         "Fruit",     2, 2, "M&S Taste Buds Seedless Easy Peelers 500g",             "", ""),
    # Tier 2 — Health
    ("Westlab Epsom Salts",               "Health",    2, 1, "Westlab Epsom Salts 2kg",                               "", ""),
    ("Haliborange Omega-3 Gummies",       "Health",    2, 1, "Haliborange Kid's Softies Omega-3 Multivitamin Gummies","", "Kids"),
    ("Bioglan SmartKids Happy Tummies",   "Health",    2, 1, "Bioglan SmartKids Vitagummies Happy Tummies 30",        "", "Kids"),
    ("Haliborange Calcium & Vit D",       "Health",    2, 1, "Haliborange Kid's Softies Calcium Vitamin D Gummies",   "", "Kids"),
    # Tier 2 — Household
    ("Laundry Soap",                      "Household", 2, 1, "Ecover Non-Bio Laundry Liquid",                         "", "⚠ Confirm product — guessed Ecover based on other household items"),
    # Tier 2 — Snacks
    ("Collective Suckies Strawberry",     "Snacks",    2, 2, "The Collective Suckies Strawberry Yoghurt 90g",         "", "Kids"),
    ("Collective Suckies Raspberry",      "Snacks",    2, 2, "The Collective Suckies Raspberry Yoghurt 90g",          "", "Kids"),
    ("Collective Suckies Peach & Apricot","Snacks",    2, 2, "The Collective Suckies Peach Apricot Yoghurt 90g",      "", "Kids"),
    ("Sliced Turkey",                     "Snacks",    2, 2, "M&S British Wafer Thin Turkey",                         "", "⚠ Confirm exact product"),
    # Tier 2 — Veg
    ("Cauliflower",                       "Veg",       2, 1, "M&S Organic Cauliflower",                               "", "⚠ Confirm organic/non-organic"),
    # Tier 3 — Pasta
    ("GI High Protein Spaghetti",         "Pasta",     3, 1, "GI Pasta High Protein Fibre Low Glycaemic Spaghetti 250g","", ""),
    ("Garofalo High Protein Mezze Maniche","Pasta",    3, 1, "Garofalo High Protein Mezze Maniche Pasta 500g",        "", ""),
    # Tier 3 — Treats
    ("Pots & Co Little Chocolate Pots",   "Treats",    3, 1, "Pots Co Little Pots of Chocolate 4x50g",               "", ""),
    ("Pots & Co Salted Caramel Ganache",  "Treats",    3, 1, "Pots Co Salted Caramel Chocolate Ganache Twin Pack",    "", ""),
    # Tier 3 — Snacks
    ("Gran Luchito Corn Tortilla Chips",  "Snacks",    3, 1, "Gran Luchito Lightly Salted Corn Tortilla Chips 170g",  "", ""),
    ("Huligan Pretzel Honey Mustard",     "Snacks",    3, 1, "Huligan Pretzel Crush Honey Mustard 65g",               "", ""),
    ("Crazy Jack Organic Cranberries",    "Snacks",    3, 1, "Crazy Jack Organic Cranberries 100g",                   "", ""),
    ("Giving Tree Freeze Dried Strawberry","Snacks",   3, 1, "Giving Tree Freeze Dried Strawberry Crisps 38g",        "", ""),
    ("BEAR Coconut Chips",                "Snacks",    3, 1, "BEAR Fruit Dried Coconut Chips 25g",                    "", ""),
    ("Nairn's Dark Choc & Orange Biscuits","Snacks",   3, 1, "Nairn's Dark Chocolate Orange Oat Biscuits 200g",       "", "6 boxes in cupboard!"),
    ("GAIL's Seeded Crackers",            "Snacks",    3, 1, "GAIL's Seeded Crackers 200g",                           "", ""),
    # Tier 3 — Household
    ("Ecover Toilet Cleaner",             "Household", 3, 1, "Ecover Toilet Cleaner Power 750ml",                     "", ""),
    ("Duck Toilet Gel",                   "Household", 3, 1, "Duck Deep Action Gel Toilet Liquid Cleaner Marine 750ml","",""),
    ("Ecozone Dishwasher Tablets",        "Household", 3, 1, "Ecozone Optimum All-in-One Dishwasher Tablets 30",      "", ""),
    ("Neat Dishwasher Tablets",           "Household", 3, 1, "Neat All in One Dishwasher Tablets Lemon 30",           "", ""),
    ("Cushelle Toilet Roll Mega 12pk",    "Household", 3, 1, "Cushelle Original Mega Toilet Rolls 12 pack",           "", ""),
    ("Ocado Compostable Caddy Liners",    "Household", 3, 1, "Ocado Compostable Caddy Liners 5L 20pk",                "", ""),
]

wb = Workbook()

# ── SHEET 1: TIER 1 ESSENTIALS ───────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Tier 1 – Essentials"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

col_widths = [38, 12, 6, 42, 32, 8]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

write_banner(ws1, 1, "🛒  OCADO WEEKLY ESSENTIALS — Added automatically every Tuesday", 6)
write_headers(ws1, 2, ["Item Name", "Category", "Qty", "Ocado Search Term", "Notes", "Active?"])

for r, row in enumerate(tier1_rows, 3):
    cat_bg = CAT_COLOURS.get(row[1], "FFFFFF")
    for i, v in enumerate(row, 1):
        c = ws1.cell(row=r, column=i, value=v)
        is_warn = isinstance(v, str) and v.startswith("⚠")
        c.fill = fill(PINK if is_warn else cat_bg)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border
    ws1.row_dimensions[r].height = 16

# ── SHEET 2: ROTATION POOL ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Rotation Pool")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

col_widths2 = [38, 12, 6, 6, 42, 14, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

write_banner(ws2, 1, "🔄  ROTATION POOL — Tier 2: check fridge first  |  Tier 3: add manually  |  ⚠ = confirm search term", 7)
write_headers(ws2, 2, ["Item Name", "Category", "Tier", "Qty", "Ocado Search Term", "Last Ordered", "Notes"])

for r, row in enumerate(rotation_rows, 3):
    tier = row[2]
    base_bg = AMBER if tier == 2 else RED_LITE
    for i, v in enumerate(row, 1):
        c = ws2.cell(row=r, column=i, value=v)
        is_warn = isinstance(v, str) and v.startswith("⚠")
        c.fill = fill(PINK if is_warn else base_bg)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border
    ws2.row_dimensions[r].height = 16

# ── SHEET 3: SETTINGS ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Settings")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 50
ws3.column_dimensions["B"].width = 20

write_banner(ws3, 1, "⚙️  SETTINGS & HOW IT WORKS", 2)

sections = [
    ("HOW IT WORKS", None),
    ("Every Tuesday evening, the automation:", None),
    ("1. Clears the upcoming Wednesday Reserved order", None),
    ("2. Adds every Active Tier 1 item automatically", None),
    ("3. Sends you a Tier 2 checklist — tap Yes/No after checking the fridge", None),
    ("4. Adds your Yes choices to the order", None),
    ("5. Tier 3 items must be added manually", None),
    ("", None),
    ("CONFIGURATION", None),
    ("Protein slots per week (from Tier 2 rotation)", 1),
    ("", None),
    ("NOTES", None),
    ("Set Active? = No on Essentials sheet to temporarily skip an item", None),
    ("Last Ordered column on Rotation Pool is updated by automation for rotation tracking", None),
    ("Items marked ⚠ need the search term confirmed before automation can add them", None),
]

for r, (label, val) in enumerate(sections, 2):
    is_section = label in ("HOW IT WORKS", "CONFIGURATION", "NOTES")
    c = ws3.cell(row=r, column=1, value=label)
    c.font = Font(name="Calibri", bold=is_section, size=10)
    c.fill = fill("E8EAF6" if is_section else WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border
    ws3.row_dimensions[r].height = 18 if is_section else 15

    c2 = ws3.cell(row=r, column=2, value=val)
    c2.font = Font(name="Calibri", size=10)
    c2.fill = fill("E8EAF6" if is_section else WHITE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = border

out = "/sessions/quirky-elegant-maxwell/mnt/outputs/Ocado_Order_Manager.xlsx"
wb.save(out)
print("Saved:", out)
