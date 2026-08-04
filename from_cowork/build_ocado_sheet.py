from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from datetime import date

wb = Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
PURPLE      = "6B3FA0"   # Ocado brand purple
PURPLE_LITE = "EDE7F6"
GREEN       = "2E7D32"
GREEN_LITE  = "E8F5E9"
AMBER_LITE  = "FFF8E1"
BLUE_LITE   = "E3F2FD"
GREY_DARK   = "424242"
GREY_LITE   = "F5F5F5"
WHITE       = "FFFFFF"
RED_LITE    = "FFEBEE"

def hdr_font(bold=True, colour=WHITE, sz=10):
    return Font(name="Arial", bold=bold, color=colour, size=sz)

def body_font(bold=False, colour=GREY_DARK, sz=10):
    return Font(name="Arial", bold=bold, color=colour, size=sz)

def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Sheet 1 : Tier 1 – Weekly Essentials ───────────────────────────────────
ws1 = wb.active
ws1.title = "Tier 1 – Essentials"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

# Title row
ws1.merge_cells("A1:F1")
ws1["A1"] = "🛒  OCADO WEEKLY ESSENTIALS  —  Added automatically every Tuesday"
ws1["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1["A1"].fill = fill(PURPLE)
ws1["A1"].alignment = centre()
ws1.row_dimensions[1].height = 28

# Header row
headers1 = ["Item Name", "Category", "Qty", "Ocado Search Term", "Notes", "Active?"]
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = hdr_font()
    cell.fill = fill(GREY_DARK)
    cell.alignment = centre()
    cell.border = thin_border()
ws1.row_dimensions[2].height = 22

essentials = [
    # Fruit
    ("M&S Pink Lady Apples", "Fruit", 1, "M&S Pink Lady Apples", "", "Yes"),
    ("M&S Organic Fairtrade Bananas", "Fruit", 1, "M&S Organic Fairtrade Bananas", "", "Yes"),
    ("M&S Blackberries", "Fruit", 1, "M&S Blackberries 150g", "", "Yes"),
    ("M&S Blueberries Family Pack", "Fruit", 2, "M&S Blueberries Family Pack 300g", "", "Yes"),
    ("M&S Raspberries", "Fruit", 1, "M&S Raspberries 250g", "", "Yes"),
    ("M&S Seedless White Grapes", "Fruit", 1, "M&S Seedless White Grapes 500g", "", "Yes"),
    # Veg
    ("M&S Extra Fine Beans", "Veg", 1, "M&S Extra Fine Beans 200g", "", "Yes"),
    ("M&S Organic Broccoli", "Veg", 1, "M&S Organic Broccoli 350g", "", "Yes"),
    ("M&S Organic Cucumber", "Veg", 1, "M&S Organic Cucumber", "", "Yes"),
    ("M&S Organic Peppers", "Veg", 1, "M&S Organic Peppers", "", "Yes"),
    ("M&S Organic Cherry Tomatoes", "Veg", 1, "M&S Organic Piccolini Cherry Tomatoes Vine Ripened", "", "Yes"),
    # Dairy
    ("Yeo Valley Organic Whole Milk", "Dairy", 1, "Yeo Valley Organic Fresh Whole Milk 1L", "", "Yes"),
    ("Moo Long Life Whole Milk", "Dairy", 1, "Moo Whole Long Life Milk 1L", "", "Yes"),
    ("M&S Mature Grated Cheddar", "Dairy", 1, "M&S British Mature Grated Cheddar 250g", "", "Yes"),
    ("M&S Double Cream", "Dairy", 1, "M&S British Double Cream 600ml", "", "Yes"),
    # Eggs
    ("Burford Brown Large Eggs x10", "Eggs", 3, "Clarence Court Burford Brown 10 Large Free Range Eggs", "", "Yes"),
    # Breakfast
    ("Cheerios Honey Multigrain", "Breakfast", 1, "Cheerios Honey Multigrain Breakfast Cereal 370g", "", "Yes"),
    ("Jordans Strawberry Country Crisp", "Breakfast", 1, "Jordans Country Crisp Sun-Ripe Strawberry 450g", "", "Yes"),
    # Pasta
    ("GI High Protein Spaghetti", "Pasta", 1, "GI Pasta High Protein Fibre Low Glycaemic Spaghetti", "", "Yes"),
    ("Garofalo High Protein Mezze Maniche", "Pasta", 1, "Garofalo High Protein Mezze Maniche Pasta", "", "Yes"),
    # Snacks
    ("Bastides Petits Saucissons Secs", "Snacks", 2, "Bastides Petits Saucissons Secs", "Buy 2 for £7 deal", "Yes"),
    ("Popz Butter Microwave Popcorn", "Snacks", 1, "Popz Butter Microwave Popcorn 6x90g", "", "Yes"),
    ("M&S Mint Crumbles", "Snacks", 2, "M&S Mint Crumbles 178g", "", "Yes"),
    # Drinks
    ("Diet Coke No Caffeine", "Drinks", 1, "Diet Coke No Caffeine 8x330ml", "", "Yes"),
    # Household
    ("Ecover Toilet Cleaner", "Household", 1, "Ecover Toilet Cleaner Power 750ml", "Check stock first", "Yes"),
    ("Duck Toilet Gel", "Household", 1, "Duck Deep Action Gel Toilet Liquid Cleaner Marine 750ml", "Check stock first", "Yes"),
    ("Ecozone Dishwasher Tablets", "Household", 1, "Ecozone Optimum All-in-One Dishwasher Tablets 30", "Check stock first", "Yes"),
    ("Neat Dishwasher Tablets", "Household", 1, "Neat All in One Dishwasher Tablets Lemon 30", "Check stock first", "Yes"),
    ("Cushelle Toilet Roll Mega 12pk", "Household", 1, "Cushelle Original Mega Toilet Rolls 12 pack", "Check stock first", "Yes"),
    ("Ocado Compostable Caddy Liners", "Household", 1, "Ocado Compostable Caddy Liners 5L 20pk", "Check stock first", "Yes"),
]

CAT_COLOURS = {
    "Fruit":      "FCE4EC",
    "Veg":        "E8F5E9",
    "Dairy":      "E3F2FD",
    "Eggs":       "FFF9C4",
    "Breakfast":  "FFF3E0",
    "Pasta":      "F3E5F5",
    "Snacks":     "FBE9E7",
    "Drinks":     "E0F7FA",
    "Household":  "EFEBE9",
}

for row_idx, row_data in enumerate(essentials, 3):
    cat = row_data[1]
    row_fill = fill(CAT_COLOURS.get(cat, "FFFFFF"))
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font()
        cell.fill = row_fill
        cell.border = thin_border()
        cell.alignment = centre() if col_idx in (2, 3, 6) else left()
    ws1.row_dimensions[row_idx].height = 18

# Column widths
for col, width in zip("ABCDEF", [38, 14, 6, 42, 28, 8]):
    ws1.column_dimensions[get_column_letter(col.index("ABCDEF".index(col)+1) if False else ord(col)-64)].width = width
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 14
ws1.column_dimensions["C"].width = 6
ws1.column_dimensions["D"].width = 42
ws1.column_dimensions["E"].width = 28
ws1.column_dimensions["F"].width = 8

# ── Sheet 2 : Rotation Pool ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Rotation Pool")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

ws2.merge_cells("A1:G1")
ws2["A1"] = "🔄  ROTATION POOL  —  Tier 2 = check fridge first  |  Tier 3 = occasional treats"
ws2["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws2["A1"].fill = fill(GREEN)
ws2["A1"].alignment = centre()
ws2.row_dimensions[1].height = 28

headers2 = ["Item Name", "Category", "Tier", "Default Qty", "Ocado Search Term", "Last Ordered", "Notes"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = hdr_font()
    cell.fill = fill(GREEN)
    cell.alignment = centre()
    cell.border = thin_border()
ws2.row_dimensions[2].height = 22

rotation_pool = [
    # Proteins
    ("M&S 2 Chicken Kyivs", "Protein", 2, 2, "M&S 2 Chicken Kyivs 320g", None, "Rotate with other proteins"),
    ("M&S Organic Chicken Drumsticks", "Protein", 2, 3, "M&S Organic British Chicken Drumsticks 500g", None, "Rotate with other proteins"),
    ("M&S Duck Breast Portions", "Protein", 2, 2, "M&S Select Farms British 2 Duck Breast Portions 265g", None, "Buy 2 for £10 deal"),
    ("M&S Scottish Salmon Fillets", "Protein", 2, 1, "M&S 4 Scottish Salmon Fillets Skin On 480g", None, "Rotate with other proteins"),
    ("Unearthed Mild Kabanos", "Protein", 2, 1, "Unearthed Mild Kabanos 105g", None, "Rotate with other proteins"),
    # Cheese
    ("Snowdonia Black Bomber Cheddar", "Cheese", 2, 1, "Snowdonia Black Bomber Extra Mature Cheddar 200g", None, ""),
    ("M&S Somerset Ripening Brie", "Cheese", 2, 1, "M&S Somerset Ripening Brie 230g", None, ""),
    ("Boursin Garlic & Herbs", "Cheese", 2, 1, "Boursin Garlic Herbs Soft French Cheese 150g", None, ""),
    ("M&S Shaved Parmigiano Reggiano", "Cheese", 2, 1, "M&S Shaved Parmigiano Reggiano 80g", None, ""),
    ("Cathedral City Mini Snack Cheeses", "Cheese", 2, 1, "Cathedral City Mini Mature Snack Cheeses 6x20g", None, "Kids snack"),
    ("Cheestrings Original", "Cheese", 2, 1, "Cheestrings Original 8x20g", None, "Kids snack"),
    # Butter / extra dairy
    ("M&S Salted Butter from Brittany", "Dairy", 2, 1, "M&S Salted Butter from Brittany 250g", None, ""),
    ("Yeo Valley Organic Unsalted Butter", "Dairy", 2, 1, "Yeo Valley Organic Unsalted Butter 200g", None, ""),
    # Fruit extra
    ("M&S Seedless Easy Peelers", "Fruit", 2, 1, "M&S Taste Buds Seedless Easy Peelers 500g", None, ""),
    # Frozen
    ("Strong Roots Sweet Potato Fries", "Frozen", 2, 1, "Strong Roots Oven Baked Sweet Potato Fries 500g", None, ""),
    ("Aunt Bessie's Roast Potatoes", "Frozen", 2, 1, "Aunt Bessie's Roast Potatoes 1.1kg", None, ""),
    # Drinks
    ("Cawston Press Cloudy Apple Juice", "Drinks", 2, 1, "Cawston Press Cloudy Apple Juice 1L", None, ""),
    ("Dalston's Sparkling Cherry", "Drinks", 2, 1, "Dalston's Sparkling Cherry 4x330ml", None, ""),
    ("London Essence Pink Grapefruit Cans", "Drinks", 2, 1, "London Essence Co Pink Grapefruit Cans 6x150ml", None, ""),
    # Health & supplements
    ("Westlab Epsom Salts", "Health", 2, 1, "Westlab Epsom Salts 2kg", None, ""),
    ("Haliborange Omega-3 Gummies", "Health", 2, 1, "Haliborange Kid's Softies Omega-3 Multivitamin Orange Gummies 3-12yrs", None, "Kids"),
    ("Bioglan SmartKids Happy Tummies", "Health", 2, 1, "Bioglan SmartKids Vitagummies Happy Tummies 30", None, "Kids"),
    ("Haliborange Calcium & Vit D Gummies", "Health", 2, 1, "Haliborange Kid's Softies Calcium Vitamin D Strawberry Gummies 3-12yrs", None, "Kids"),
    # Occasional treats
    ("Pots & Co Little Chocolate Pots", "Treats", 3, 1, "Pots Co Little Pots of Chocolate 4x50g", None, ""),
    ("Pots & Co Salted Caramel Ganache", "Treats", 3, 1, "Pots Co Salted Caramel Chocolate Ganache Twin Pack", None, ""),
    ("Gran Luchito Corn Tortilla Chips", "Snacks", 3, 1, "Gran Luchito Lightly Salted Corn Tortilla Chips 170g", None, ""),
    ("Huligan Pretzel Crush Honey Mustard", "Snacks", 3, 1, "Huligan Pretzel Crush Honey Mustard 65g", None, ""),
    ("Crazy Jack Organic Cranberries", "Snacks", 3, 1, "Crazy Jack Organic Cranberries 100g", None, ""),
    ("Giving Tree Freeze Dried Strawberry", "Snacks", 3, 1, "Giving Tree Freeze Dried Strawberry Crisps 38g", None, ""),
    ("BEAR Coconut Chips", "Snacks", 3, 1, "BEAR Fruit Dried Coconut Chips 25g", None, ""),
    ("Nairn's Dark Choc & Orange Biscuits", "Snacks", 3, 1, "Nairn's Dark Chocolate Orange Oat Biscuits 200g", None, "6 boxes in cupboard!"),
    ("GAIL's Seeded Crackers", "Snacks", 3, 1, "GAIL's Seeded Crackers 200g", None, ""),
]

TIER_COLOURS = {2: AMBER_LITE, 3: RED_LITE}

for row_idx, row_data in enumerate(rotation_pool, 3):
    tier = row_data[2]
    row_fill = fill(TIER_COLOURS.get(tier, WHITE))
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font()
        cell.fill = row_fill
        cell.border = thin_border()
        if col_idx == 6 and value is None:
            cell.value = ""
            cell.number_format = "DD/MM/YYYY"
        cell.alignment = centre() if col_idx in (2, 3, 4, 6) else left()
    ws2.row_dimensions[row_idx].height = 18

ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 6
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 48
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 28

# ── Sheet 3 : Settings & How It Works ───────────────────────────────────────
ws3 = wb.create_sheet("Settings")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
ws3["A1"] = "⚙️  SETTINGS & HOW IT WORKS"
ws3["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws3["A1"].fill = fill(GREY_DARK)
ws3["A1"].alignment = centre()
ws3.row_dimensions[1].height = 28

settings_data = [
    ("HOW IT WORKS", None, None),
    ("Every Tuesday evening, the automation:", None, None),
    ("  1. Clears the upcoming Wednesday Reserved order", None, None),
    ("  2. Adds every active Tier 1 item (Essentials sheet) to the order", None, None),
    ("  3. Presents the Tier 2 checklist for you to review (fridge check!)", None, None),
    ("  4. Adds any ticked Tier 2 items to the order", None, None),
    ("  5. Tier 3 items must be added manually if wanted", None, None),
    (None, None, None),
    ("ROTATION SETTINGS", None, None),
    ("Protein picks per week", "Protein slots", 1),
    ("Max snack variety items per week (Tier 3)", "Snack slots", 2),
    (None, None, None),
    ("NOTES", None, None),
    ("Tier 1 = add every week, no questions asked", None, None),
    ("Tier 2 = check stock first — automation will prompt you on Tuesday", None, None),
    ("Tier 3 = occasional / one-offs — add manually when wanted", None, None),
    ("Last Ordered column in Rotation Pool is updated by the automation", None, None),
    ("Set Active? = No on Tier 1 Essentials sheet to skip an item temporarily", None, None),
]

for row_idx, (a, b, c) in enumerate(settings_data, 2):
    ws3.row_dimensions[row_idx].height = 18
    if a and a == a.upper() and b is None:
        ws3.merge_cells(f"A{row_idx}:C{row_idx}")
        cell = ws3.cell(row=row_idx, column=1, value=a)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = fill(PURPLE)
        cell.alignment = left()
    elif a and a.startswith("  "):
        cell = ws3.cell(row=row_idx, column=1, value=a)
        cell.font = body_font()
        cell.alignment = left()
    elif b is not None:
        ws3.cell(row=row_idx, column=1, value=a).font = body_font(bold=True)
        ws3.cell(row=row_idx, column=2, value=b).font = body_font()
        cell = ws3.cell(row=row_idx, column=3, value=c)
        cell.font = Font(name="Arial", bold=True, color="0000FF", size=10)
        cell.fill = fill("FFFDE7")
        cell.alignment = centre()
        cell.border = thin_border()
    elif a:
        cell = ws3.cell(row=row_idx, column=1, value=a)
        cell.font = body_font()
        cell.alignment = left()
        ws3.merge_cells(f"A{row_idx}:C{row_idx}")

ws3.column_dimensions["A"].width = 52
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 12

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = "/sessions/quirky-elegant-maxwell/mnt/outputs/Ocado_Order_Manager.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
