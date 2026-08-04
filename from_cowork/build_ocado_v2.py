from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

PURPLE    = "6B3FA0"
GREEN     = "2E7D32"
TEAL      = "00695C"
GREY_DARK = "424242"
WHITE     = "FFFFFF"
AMBER     = "FFF8E1"
RED_LITE  = "FFEBEE"
NEEDS_REV = "FCE4EC"

CAT_COLOURS = {
    "Breakfast":   "FFF3E0",
    "Cheese":      "FFF9C4",
    "Condiments":  "F3E5F5",
    "Dairy":       "E3F2FD",
    "Drinks":      "E0F7FA",
    "Eggs":        "FFFDE7",
    "Fruit":       "FCE4EC",
    "Household":   "EFEBE9",
    "Snacks":      "FBE9E7",
    "Veg":         "E8F5E9",
}

def fill(h): return PatternFill("solid", fgColor=h)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def hdr_font(c=WHITE): return Font(name="Arial", bold=True, color=c, size=10)
def body_font(bold=False, c=GREY_DARK): return Font(name="Arial", bold=bold, color=c, size=10)
def centre(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_title(ws, text, colour, cols="A1:F1"):
    ws.merge_cells(cols)
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
    ws["A1"].fill = fill(colour)
    ws["A1"].alignment = centre()
    ws.row_dimensions[1].height = 28

def write_headers(ws, headers, colour, row=2):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font()
        c.fill = fill(colour)
        c.alignment = centre()
        c.border = thin_border()
    ws.row_dimensions[row].height = 22

def write_rows(ws, rows, start=3, tier_col=None):
    for i, row in enumerate(rows):
        r = start + i
        cat = row[1] if len(row) > 1 else ""
        tier = row[2] if tier_col and len(row) > tier_col else None
        if row[0].startswith("⚠️"):
            bg = NEEDS_REV
        elif tier == 3:
            bg = RED_LITE
        elif tier == 2:
            bg = AMBER
        else:
            bg = CAT_COLOURS.get(cat, WHITE)
        for j, val in enumerate(row):
            c = ws.cell(row=r, column=j+1, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = centre() if j in (1,2,3,5) else left()
        ws.row_dimensions[r].height = 18

# ── SHEET 1: Tier 1 Essentials ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Tier 1 – Essentials"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

write_title(ws1, "🛒  OCADO WEEKLY ESSENTIALS  —  Added automatically every Tuesday", PURPLE)
write_headers(ws1, ["Item Name", "Category", "Qty", "Ocado Search Term", "Notes", "Active?"], GREY_DARK)

essentials = [
    # Fruit
    ("M&S Pink Lady Apples",          "Fruit",      1, "M&S Pink Lady Apples",                                    "", "Yes"),
    ("M&S Organic Fairtrade Bananas", "Fruit",      1, "M&S Organic Fairtrade Bananas",                           "", "Yes"),
    ("M&S Blueberries Family Pack",   "Fruit",      2, "M&S Blueberries Family Pack 300g",                        "", "Yes"),
    ("M&S Raspberries",               "Fruit",      1, "M&S Raspberries 250g",                                    "", "Yes"),
    ("M&S Blackberries",              "Fruit",      1, "M&S Blackberries 150g",                                   "", "Yes"),
    ("M&S Seedless White Grapes",     "Fruit",      1, "M&S Seedless White Grapes 500g",                          "", "Yes"),
    ("Oranges",                       "Fruit",      1, "oranges",                                                  "⚠ Add proper search term", "Yes"),
    # Veg
    ("M&S Extra Fine Beans",          "Veg",        1, "M&S Extra Fine Beans 200g",                               "", "Yes"),
    ("M&S Organic Broccoli",          "Veg",        1, "M&S Organic Broccoli 350g",                               "", "Yes"),
    ("M&S Organic Cucumber",          "Veg",        1, "M&S Organic Cucumber",                                    "", "Yes"),
    ("M&S Organic Peppers",           "Veg",        1, "M&S Organic Peppers",                                     "", "Yes"),
    ("M&S Organic Cherry Tomatoes",   "Veg",        1, "M&S Organic Piccolini Cherry Tomatoes Vine Ripened",      "", "Yes"),
    # Dairy
    ("Yeo Valley Organic Whole Milk", "Dairy",      1, "Yeo Valley Organic Fresh Whole Milk 1L",                  "", "Yes"),
    ("Moo Long Life Whole Milk",      "Dairy",      1, "Moo Whole Long Life Milk 1L",                             "", "Yes"),
    ("M&S Mature Grated Cheddar",     "Dairy",      1, "M&S British Mature Grated Cheddar 250g",                  "", "Yes"),
    ("M&S Double Cream",              "Dairy",      1, "M&S British Double Cream 600ml",                          "", "Yes"),
    # Cheese
    ("Gouda",                         "Cheese",     1, "gouda",                                                    "⚠ Add proper search term", "Yes"),
    # Eggs
    ("Burford Brown Large Eggs x10",  "Eggs",       3, "Clarence Court Burford Brown 10 Large Free Range Eggs",   "", "Yes"),
    # Breakfast
    ("Cheerios Honey Multigrain",         "Breakfast", 1, "Cheerios Honey Multigrain Breakfast Cereal 370g",      "", "Yes"),
    ("Jordans Strawberry Country Crisp",  "Breakfast", 1, "Jordans Country Crisp Sun-Ripe Strawberry 450g",       "", "Yes"),
    ("Protein Everything Bagels",         "Breakfast", 1, "protein everything bagels",                             "⚠ Add proper search term", "Yes"),
    ("Oatly Barista (grey)",              "Breakfast", 1, "Oatly Barista Edition oat drink",                      "", "Yes"),
    # Condiments
    ("Ketchup",                       "Condiments", 1, "ketchup",                                                  "⚠ Add preferred brand", "Yes"),
    ("Mayo",                          "Condiments", 1, "mayonnaise",                                               "⚠ Add preferred brand", "Yes"),
    ("Maple Syrup",                   "Condiments", 1, "maple syrup",                                              "⚠ Add preferred brand", "Yes"),
    # Drinks
    ("Diet Coke No Caffeine",         "Drinks",     1, "Diet Coke No Caffeine 8x330ml",                           "", "Yes"),
    # Snacks
    ("Bastides Petits Saucissons Secs","Snacks",    2, "Bastides Petits Saucissons Secs",                         "Buy 2 for £7 deal", "Yes"),
    ("Popz Butter Microwave Popcorn", "Snacks",     1, "Popz Butter Microwave Popcorn 6x90g",                     "", "Yes"),
    ("M&S Mint Crumbles",             "Snacks",     2, "M&S Mint Crumbles 178g",                                  "", "Yes"),
    # Household
    ("Ecover Toilet Cleaner",         "Household",  1, "Ecover Toilet Cleaner Power 750ml",                       "Check stock first", "Yes"),
    ("Duck Toilet Gel",               "Household",  1, "Duck Deep Action Gel Toilet Liquid Cleaner Marine 750ml", "Check stock first", "Yes"),
    ("Ecozone Dishwasher Tablets",    "Household",  1, "Ecozone Optimum All-in-One Dishwasher Tablets 30",        "Check stock first", "Yes"),
    ("Neat Dishwasher Tablets",       "Household",  1, "Neat All in One Dishwasher Tablets Lemon 30",             "Check stock first", "Yes"),
    ("Cushelle Toilet Roll Mega 12pk","Household",  1, "Cushelle Original Mega Toilet Rolls 12 pack",             "Check stock first", "Yes"),
    ("Ocado Compostable Caddy Liners","Household",  1, "Ocado Compostable Caddy Liners 5L 20pk",                  "Check stock first", "Yes"),
]

write_rows(ws1, essentials)

ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 13
ws1.column_dimensions["C"].width = 6
ws1.column_dimensions["D"].width = 48
ws1.column_dimensions["E"].width = 26
ws1.column_dimensions["F"].width = 8

# ── SHEET 2: Rotation Pool ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Rotation Pool")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

write_title(ws2, "🔄  ROTATION POOL  —  Tier 2 = check first  |  Tier 3 = occasional  |  ⚠️ = needs search term", GREEN)
write_headers(ws2, ["Item Name", "Category", "Tier", "Default Qty", "Ocado Search Term", "Last Ordered", "Notes"], GREEN)

rotation = [
    # ── TIER 2 ──
    # Protein
    ("M&S 2 Chicken Kyivs",               "Protein", 2, 2, "M&S 2 Chicken Kyivs 320g",                                    "", "Rotate with other proteins"),
    ("M&S Organic Chicken Drumsticks",     "Protein", 2, 3, "M&S Organic British Chicken Drumsticks 500g",                 "", "Rotate with other proteins"),
    ("M&S Duck Breast Portions",           "Protein", 2, 2, "M&S Select Farms British 2 Duck Breast Portions 265g",        "", "Buy 2 for £10 deal"),
    ("M&S Scottish Salmon Fillets",        "Protein", 2, 1, "M&S 4 Scottish Salmon Fillets Skin On 480g",                  "", "Rotate with other proteins"),
    ("Unearthed Mild Kabanos",             "Protein", 2, 1, "Unearthed Mild Kabanos 105g",                                 "", ""),
    ("⚠️ Pork Chops",                      "Protein", 2, 1, "",                                                             "", "Needs search term — which brand/cut?"),
    ("⚠️ Turkey Steaks",                   "Protein", 2, 1, "",                                                             "", "Needs search term — which brand/cut?"),
    # Cheese
    ("Snowdonia Black Bomber Cheddar",     "Cheese",  2, 1, "Snowdonia Black Bomber Extra Mature Cheddar 200g",            "", ""),
    ("M&S Somerset Ripening Brie",         "Cheese",  2, 1, "M&S Somerset Ripening Brie 230g",                             "", ""),
    ("Boursin Garlic & Herbs",             "Cheese",  2, 1, "Boursin Garlic Herbs Soft French Cheese 150g",                "", ""),
    ("M&S Shaved Parmigiano Reggiano",     "Cheese",  2, 1, "M&S Shaved Parmigiano Reggiano 80g",                          "", ""),
    ("Cathedral City Mini Snack Cheeses",  "Cheese",  2, 1, "Cathedral City Mini Mature Snack Cheeses 6x20g",              "", "Kids snack"),
    ("Cheestrings Original",               "Cheese",  2, 1, "Cheestrings Original 8x20g",                                  "", "Kids snack"),
    ("Leerdammer",                         "Cheese",  2, 2, "Leerdammer",                                                   "", ""),
    # Dairy
    ("M&S Salted Butter from Brittany",    "Dairy",   2, 1, "M&S Salted Butter from Brittany 250g",                        "", ""),
    ("Yeo Valley Organic Unsalted Butter", "Dairy",   2, 1, "Yeo Valley Organic Unsalted Butter 200g",                     "", ""),
    # Drinks
    ("Cawston Press Cloudy Apple Juice",   "Drinks",  2, 1, "Cawston Press Cloudy Apple Juice 1L",                         "", ""),
    ("Dalston's Sparkling Cherry",         "Drinks",  2, 1, "Dalston's Sparkling Cherry 4x330ml",                          "", ""),
    ("London Essence Pink Grapefruit",     "Drinks",  2, 1, "London Essence Co Pink Grapefruit Cans 6x150ml",              "", ""),
    ("⚠️ Caffeinated Diet Coke/Pepsi",     "Drinks",  2, 1, "",                                                             "", "Needs search term — Diet Coke / Pepsi Max?"),
    # Frozen
    ("Strong Roots Sweet Potato Fries",    "Frozen",  2, 1, "Strong Roots Oven Baked Sweet Potato Fries 500g",             "", ""),
    ("Aunt Bessie's Roast Potatoes",       "Frozen",  2, 1, "Aunt Bessie's Roast Potatoes 1.1kg",                          "", ""),
    # Fruit
    ("M&S Seedless Easy Peelers",          "Fruit",   2, 2, "M&S Taste Buds Seedless Easy Peelers 500g",                   "", ""),
    # Health
    ("Westlab Epsom Salts",                "Health",  2, 1, "Westlab Epsom Salts 2kg",                                     "", ""),
    ("Haliborange Omega-3 Gummies",        "Health",  2, 1, "Haliborange Kid's Softies Omega-3 Multivitamin Gummies",      "", "Kids"),
    ("Bioglan SmartKids Happy Tummies",    "Health",  2, 1, "Bioglan SmartKids Vitagummies Happy Tummies 30",              "", "Kids"),
    ("Haliborange Calcium & Vit D",        "Health",  2, 1, "Haliborange Kid's Softies Calcium Vitamin D Gummies",         "", "Kids"),
    # Household
    ("⚠️ Laundry Soap",                    "Household",2,1, "",                                                             "", "Needs search term — which product?"),
    # Snacks / Yoghurts
    ("Collective Suckies Strawberry",      "Snacks",  2, 2, "The Collective Suckies Strawberry Yoghurt 90g",               "", "Kids"),
    ("Collective Suckies Raspberry",       "Snacks",  2, 2, "The Collective Suckies Raspberry Yoghurt 90g",                "", "Kids"),
    ("Collective Suckies Peach & Apricot", "Snacks",  2, 2, "The Collective Suckies Peach Apricot Yoghurt 90g",            "", "Kids"),
    ("Sliced Turkey",                      "Snacks",  2, 2, "sliced turkey",                                                "", "⚠ Add preferred brand"),
    # Veg
    ("⚠️ Cauliflower",                     "Veg",     2, 1, "",                                                             "", "Needs search term — organic? M&S?"),
    # ── TIER 3 ──
    # Pasta
    ("GI High Protein Spaghetti",          "Pasta",   3, 1, "GI Pasta High Protein Fibre Low Glycaemic Spaghetti",         "", ""),
    ("Garofalo High Protein Mezze Maniche","Pasta",   3, 1, "Garofalo High Protein Mezze Maniche Pasta",                   "", ""),
    # Treats
    ("Pots & Co Little Chocolate Pots",    "Treats",  3, 1, "Pots Co Little Pots of Chocolate 4x50g",                     "", ""),
    ("Pots & Co Salted Caramel Ganache",   "Treats",  3, 1, "Pots Co Salted Caramel Chocolate Ganache Twin Pack",          "", ""),
    # Snacks
    ("Gran Luchito Corn Tortilla Chips",   "Snacks",  3, 1, "Gran Luchito Lightly Salted Corn Tortilla Chips 170g",        "", ""),
    ("Huligan Pretzel Crush Honey Mustard","Snacks",  3, 1, "Huligan Pretzel Crush Honey Mustard 65g",                     "", ""),
    ("Crazy Jack Organic Cranberries",     "Snacks",  3, 1, "Crazy Jack Organic Cranberries 100g",                         "", ""),
    ("Giving Tree Freeze Dried Strawberry","Snacks",  3, 1, "Giving Tree Freeze Dried Strawberry Crisps 38g",              "", ""),
    ("BEAR Coconut Chips",                 "Snacks",  3, 1, "BEAR Fruit Dried Coconut Chips 25g",                          "", ""),
    ("Nairn's Dark Choc & Orange Biscuits","Snacks",  3, 1, "Nairn's Dark Chocolate Orange Oat Biscuits 200g",             "", "6 boxes in cupboard!"),
    ("GAIL's Seeded Crackers",             "Snacks",  3, 1, "GAIL's Seeded Crackers 200g",                                 "", ""),
]

write_rows(ws2, rotation, tier_col=2)

ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 6
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 50
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 30

# ── SHEET 3: Settings ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Settings")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:C1")
ws3["A1"] = "⚙️  SETTINGS & HOW IT WORKS"
ws3["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws3["A1"].fill = fill(GREY_DARK)
ws3["A1"].alignment = centre()
ws3.row_dimensions[1].height = 28

settings = [
    ("HOW IT WORKS", None, None),
    ("Every Tuesday evening, the automation:", None, None),
    ("  1. Clears the upcoming Wednesday Reserved order", None, None),
    ("  2. Adds every active Tier 1 item (Essentials sheet) automatically", None, None),
    ("  3. Sends you a Tier 2 checklist — you tick what you need after checking the fridge", None, None),
    ("  4. Adds ticked Tier 2 items to the order", None, None),
    ("  5. Tier 3 items must be added manually when wanted", None, None),
    (None, None, None),
    ("CONFIGURATION", None, None),
    ("Protein picks per week (from Tier 2 Protein rotation)", "Protein slots", 1),
    ("Max Tier 3 snack items added per week", "Snack slots", 0),
    (None, None, None),
    ("NOTES", None, None),
    ("Set Active? = No on Essentials sheet to skip an item temporarily", None, None),
    ("Last Ordered column on Rotation Pool is updated by automation for rotation tracking", None, None),
    ("Items marked ⚠️ need a proper Ocado search term before automation can add them", None, None),
]

for i, (a, b, c) in enumerate(settings, 2):
    ws3.row_dimensions[i].height = 18
    if a and a == a.upper() and b is None:
        ws3.merge_cells(f"A{i}:C{i}")
        cell = ws3.cell(row=i, column=1, value=a)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = fill(PURPLE)
        cell.alignment = left()
    elif b is not None:
        ws3.cell(row=i, column=1, value=a).font = body_font(bold=True)
        ws3.cell(row=i, column=2, value=b).font = body_font()
        cell = ws3.cell(row=i, column=3, value=c)
        cell.font = Font(name="Arial", bold=True, color="0000FF", size=10)
        cell.fill = fill("FFFDE7")
        cell.alignment = centre()
        cell.border = thin_border()
    elif a:
        ws3.merge_cells(f"A{i}:C{i}")
        cell = ws3.cell(row=i, column=1, value=a)
        cell.font = body_font()
        cell.alignment = left()

ws3.column_dimensions["A"].width = 60
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 12

out = "/sessions/quirky-elegant-maxwell/mnt/outputs/Ocado_Order_Manager.xlsx"
wb.save(out)
print("Done:", out)
